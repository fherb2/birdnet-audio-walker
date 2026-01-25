"""
Database Overview Page - Database selection, metadata, and file list
"""

import streamlit as st
from pathlib import Path
from datetime import datetime, date, time as dt_time
import pandas as pd
from loguru import logger

from shared.streamlit_utils import find_databases_recursive
from shared.db_queries import get_analysis_config, get_all_metadata, set_analysis_config


st.set_page_config(
    page_title="Database Overview - BirdNET",
    page_icon="📂",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Page title
st.title("🐦 BirdNET Audio Player")
st.header("📂 Database Overview")
# Show read-only banner
if st.session_state.get('read_only', False):
    st.warning("🔒 **Read-Only Mode** - Database modifications are disabled")
st.divider()

# Database selector
st.subheader("Select Database")

# Find all databases if root_path is set
if 'root_path' in st.session_state and st.session_state['root_path']:
    available_dbs = find_databases_recursive(st.session_state['root_path'])
    
    if not available_dbs:
        st.error(f"❌ No databases found in: {st.session_state['root_path']}")
        st.stop()
    
    st.info(f"Found {len(available_dbs)} database(s)")
    
    # Selectbox for database selection (FULL WIDTH, no sidebar)
    db_options = [str(db) for db in available_dbs]
    
    # Determine current selection index
    current_index = 0
    if 'db_path' in st.session_state and st.session_state['db_path'] in available_dbs:
        current_index = available_dbs.index(st.session_state['db_path'])
    
    selected_db_str = st.selectbox(
        "Database",
        options=db_options,
        index=current_index,
        help="Choose database to analyze",
        key="db_selector_overview"
    )
    
    selected_db = Path(selected_db_str)
    
    # Check if selection changed
    if selected_db != st.session_state.get('db_path'):
        # Clear all cached data
        for key in list(st.session_state.keys()):
            if key not in ['db_path', 'root_path']:
                del st.session_state[key]
        
        # Set new database
        st.session_state['db_path'] = selected_db
        st.success(f"✅ Switched to: {selected_db.name}")
        st.rerun()
    
    db_path = selected_db
else:
    db_path = st.session_state.get('db_path')

# Check if database is valid
if db_path is None:
    st.error("❌ No database path provided")
    st.info("Start with: streamlit run streamlit_app.py -- /path/to/db.db")
    st.stop()

if not db_path.exists():
    st.error(f"❌ Database not found: {db_path}")
    st.stop()
    
# Auto-create species_list table if it doesn't exist
from shared.db_queries import species_list_exists, create_species_list_table

if not species_list_exists(db_path):
    logger.info(f"species_list table not found in {db_path}, creating...")
    if create_species_list_table(db_path):
        logger.info("species_list table created successfully")
    else:
        logger.error("Failed to create species_list table")

st.divider()

# Load database info
language_code = get_analysis_config(db_path, 'local_name_shortcut')
confidence_threshold = get_analysis_config(db_path, 'confidence_threshold')
created_at = get_analysis_config(db_path, 'created_at')

# Database Information
st.subheader("📊 Database Information")

col1, col2, col3 = st.columns(3)

with col1:
    st.metric("Language", language_code or "Unknown")

with col2:
    st.metric("Confidence Threshold", f"{float(confidence_threshold or 0)*100:.0f}%" if confidence_threshold else "Unknown")

with col3:
    st.metric("Created", created_at or "Unknown")
    
# Species count + actualize button
from shared.db_queries import get_species_count

col1, col2 = st.columns([3, 1])

with col1:
    species_count = get_species_count(db_path)
    st.metric("Number of Species", species_count if species_count > 0 else "Not available")

with col2:
    st.write("")  # Spacing
    read_only = st.session_state.get('read_only', False)

    if st.button("🔄 Actualize Species List", width="stretch", disabled=read_only):
        with st.spinner("Updating species list..."):
            if create_species_list_table(db_path):
                new_count = get_species_count(db_path)
                st.success(f"✅ Species list updated! {new_count} species found.")
                st.rerun()
            else:
                st.error("❌ Failed to update species list")
                

st.divider()

# User Comment Section
st.subheader("📝 Notes")

# Load existing comment
existing_comment = get_analysis_config(db_path, 'user_comment') or ""

# Text area for comment
user_comment = st.text_area(
    "Database Notes/Comments",
    value=existing_comment,
    height=200,
    help="Add notes or comments about this recording session",
    key=f"user_comment_{hash(str(db_path))}"
)

# Save button
if st.button("💾 Save Notes", disabled=read_only):
    success = set_analysis_config(db_path, 'user_comment', user_comment)
    if success:
        st.success("✅ Notes saved successfully!")
    else:
        st.error("❌ Failed to save notes")

st.divider()

# File List
st.subheader("📁 Recording Files")

metadata_list = get_all_metadata(db_path)

if not metadata_list:
    st.warning("⚠️ No files found in database")
    st.stop()

st.info(f"Total files: {len(metadata_list)}")

# Prepare table data
table_data = []
for meta in metadata_list:
    # Parse timestamps
    start_time = meta['timestamp_local']
    duration = meta['duration_seconds']
    
    # Calculate end time
    try:
        start_dt = datetime.fromisoformat(start_time)
        end_dt = start_dt + pd.Timedelta(seconds=duration)
        end_time = end_dt.strftime('%Y-%m-%d %H:%M:%S')
    except:
        end_time = "N/A"
    
    table_data.append({
        'Filename': meta['filename'],
        'Start Time': start_time,
        'End Time': end_time,
        'Duration (s)': f"{duration:.1f}",
        'Temperature (°C)': f"{meta['temperature_c']:.1f}" if meta['temperature_c'] else "N/A",
        'Battery (V)': f"{meta['battery_voltage']:.2f}" if meta['battery_voltage'] else "N/A",
        'GPS': f"{meta['gps_lat']:.4f}, {meta['gps_lon']:.4f}" if meta['gps_lat'] and meta['gps_lon'] else "N/A"
    })

df = pd.DataFrame(table_data)

# Display table
st.dataframe(df, width='stretch', hide_index=True)

st.divider()

# Navigation hint
st.info("➡️ Go to **Audio Player** page to filter and play detections")

st.divider()

# Species List
st.subheader("🦜 Species List")

from shared.db_queries import get_species_list_with_counts, format_detections_column
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
import pandas as pd

species_list = get_species_list_with_counts(db_path)

if not species_list:
    st.warning("⚠️ Species list not available. Click 'Actualize Species List' above.")
else:
    st.info(f"Total species: {len(species_list)}")
    
    # Prepare DataFrame
    df_species = pd.DataFrame(species_list)
    
    # Get min_score for formatting
    min_score = df_species['score'].min()
    
    # Format Detections column
    df_species['Detections'] = df_species.apply(
        lambda row: format_detections_column(
            row['count_high'], 
            row['count_low'], 
            row['score'], 
            min_score
        ), 
        axis=1
    )
    
    # Prepare for display (exclude raw count columns)
    df_display = df_species[[
        'scientific_name', 
        'local_name', 
        'name_cs', 
        'Detections',
        'score'  # Keep for sorting (will be hidden)
    ]].rename(columns={
        'scientific_name': 'Scientific Name',
        'local_name': 'Local Name',
        'name_cs': 'Czech Name'
    })
    
    # Info hint
    st.info("👆 Click on a species to select for playing - selection appears in sidebar")

    st.write("")  # Small spacing
    
    # Build grid options
    gb = GridOptionsBuilder.from_dataframe(df_display)
    
    # Configure columns
    gb.configure_column("Scientific Name", width=250)
    gb.configure_column("Local Name", width=250)
    gb.configure_column("Czech Name", width=250)
    gb.configure_column(
        "Detections", 
        width=220,
        headerName="Detections: number > 70% confidence (all) {score for sorting}",
        wrapHeaderText=True,
        autoHeaderHeight=True
    )
    gb.configure_column("score", hide=True)  # Hidden but used for sorting
    
    # Configure sorting
    gb.configure_default_column(
        sortable=True,
        resizable=True,
        filterable=True,
        suppressMenu=False
    )
    
    # Initial sort by score (descending - best species first)
    gb.configure_column(
        "score",
        sort="desc"
    )
    
    # Selection mode
    gb.configure_selection('single', use_checkbox=False, pre_selected_rows=[])
    
    # Grid options
    grid_options = gb.build()
    
    # Render grid
    grid_response = AgGrid(
        df_display,
        gridOptions=grid_options,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
        height=600,
        theme='streamlit',
        allow_unsafe_jscode=True,
        enable_enterprise_modules=False
    )
    
    # Handle selection - show in sidebar
    selected = grid_response.get('selected_rows', [])

    if selected and len(selected) > 0:
        selected_species = selected[0]['Scientific Name']
        
        with st.sidebar:
            st.divider()
            st.subheader("🎯 Selected Species")
            st.markdown(f"**{selected_species}**")
            
            if st.button("▶ Play Species", use_container_width=True, key="play_species_btn"):
                # Get date range from database
                from shared.db_queries import get_recording_date_range
                min_date, max_date = get_recording_date_range(db_path)
                
                # Reset ALL filters and set optimal defaults for species playback
                st.session_state['filter_species'] = selected_species
                st.session_state['filter_date_from'] = min_date.date() if min_date else None
                st.session_state['filter_date_to'] = max_date.date() if max_date else None
                st.session_state['filter_use_time'] = False
                st.session_state['filter_time_start'] = dt_time(0, 0)
                st.session_state['filter_time_end'] = dt_time(23, 59)
                st.session_state['filter_confidence'] = "All"
                st.session_state['filter_limit'] = 25
                st.session_state['filter_offset'] = 0
                st.session_state['filter_sort'] = "confidence"  # Sort by confidence for species view
                st.session_state['filters_applied'] = True
                
                # Switch to audio player page
                st.switch_page("pages/2_audio_player.py")
                
# Download section
st.divider()
st.subheader("📥 Download Species List")

# Radio for download scope
download_scope = st.radio(
    "Download scope:",
    options=["All species", "Only filtered species"],
    horizontal=True,
    key="download_scope_species"
)

# Prepare data based on selection
if download_scope == "Only filtered species":
    # Get filtered data from grid
    download_df = pd.DataFrame(grid_response['data'])
    # Remove score column (was hidden)
    if 'score' in download_df.columns:
        download_df = download_df.drop(columns=['score'])
    selections_text = "filtered only"
else:
    # Use original data (without score)
    download_df = df_display.drop(columns=['score'])
    selections_text = "all"

# Get date range and notes for metadata
from shared.db_queries import get_recording_date_range
min_date, max_date = get_recording_date_range(db_path)
date_from_str = min_date.strftime('%Y-%m-%d') if min_date else "N/A"
date_to_str = max_date.strftime('%Y-%m-%d') if max_date else "N/A"

user_notes = get_analysis_config(db_path, 'user_comment') or ""

col1, col2 = st.columns(2)

with col1:
    # CSV Download with header
    from io import StringIO
    
    csv_buffer = StringIO()
    
    # Write header lines
    csv_buffer.write(f"File: {db_path}\n")
    csv_buffer.write(f"Selections: {selections_text}\n")
    csv_buffer.write(f"From: {date_from_str} To: {date_to_str}\n")
    csv_buffer.write("\n")  # Empty line
    
    # Write table
    download_df.to_csv(csv_buffer, index=False)
    
    csv_data = csv_buffer.getvalue().encode('utf-8')
    
    st.download_button(
        label="📥 Download CSV",
        data=csv_data,
        file_name=f"species_list_{db_path.stem}.csv",
        mime="text/csv",
        use_container_width=True
    )

with col2:
    # Excel Download with 2 sheets
    from io import BytesIO
    from openpyxl.utils import get_column_letter
    
    buffer = BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Sheet 1: Database Info
        db_info_data = {
            'Info': [
                f"File: {db_path}",
                f"Selections: {selections_text}",
                f"From: {date_from_str} To: {date_to_str}",
                "",
                "Notes:"
            ]
        }
        
        # Add notes line by line
        if user_notes:
            for line in user_notes.split('\n'):
                db_info_data['Info'].append(line)
        
        df_info = pd.DataFrame(db_info_data)
        df_info.to_excel(writer, sheet_name='Database', index=False, header=False)
        
        # Set column width for Database sheet
        worksheet = writer.sheets['Database']
        worksheet.column_dimensions['A'].width = 80
        
        # Enable text wrapping for notes cells (from row 6 onwards)
        from openpyxl.styles import Alignment, Border, Side
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
        
        # Sheet 2: Species Table
        download_df.to_excel(writer, sheet_name='Species Table', index=False)
        
        # Set column widths for Species Table
        worksheet_species = writer.sheets['Species Table']
        worksheet_species.column_dimensions['A'].width = 19  # Scientific Name
        worksheet_species.column_dimensions['B'].width = 22  # Local Name
        worksheet_species.column_dimensions['C'].width = 17  # Czech Name
        worksheet_species.column_dimensions['D'].width = 26  # Detections
        
        # Enable text wrapping and bottom border for entire Species Table
        thin_border = Border(bottom=Side(style='thin'))
        
        for row in worksheet_species.iter_rows(min_row=1, max_row=worksheet_species.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = thin_border
                
        # Header and footer for Species Table
        worksheet_species.oddHeader.left.text = f"{db_path}"
        worksheet_species.oddFooter.right.text = "Page &P"
    
    excel_data = buffer.getvalue()
    
    st.download_button(
        label="📥 Download Excel",
        data=excel_data,
        file_name=f"species_list_{db_path.stem}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

st.divider()


