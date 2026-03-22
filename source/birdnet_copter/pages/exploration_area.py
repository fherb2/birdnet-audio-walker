"""
Exploration Area page – database selection and analysis overview.

Route: /exploration

Sections:
  1. DB selection      (DbFolderTree – multi-select checkboxes)
  2. Aggregate stats   (counts from temp_db: detections, species, files, duration)
  3. Notes             (disabled – single-DB management not yet implemented)
  4. Recording Files   (all files across all selected source DBs)
  5. Species List      (AG Grid from temp_db) + Download
"""

import asyncio
import sqlite3
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Optional, Set

import pandas as pd
from loguru import logger
from nicegui import ui, app as nicegui_app, context

from ..app_state import AppState
from ..pages.layout import create_layout
from ..gui_elements.page_header import page_header
from ..gui_elements.section_card import section_card
from ..gui_elements.db_folder_tree import DbFolderTree
from ..db_queries import (
    get_species_list_with_counts,
    format_detections_column,
    species_list_exists,
)
from ..bird_language import load_labels
from ..task_status import run_with_loading


# ---------------------------------------------------------------------------
# Page registration
# ---------------------------------------------------------------------------

def _get_state() -> AppState:
    return nicegui_app.state.app_state  # type: ignore[attr-defined]


def _get_bundle():
    return nicegui_app.state.bundle


@ui.page('/exploration')
async def exploration_area() -> None:
    state = _get_state()
    bundle = _get_bundle()
    create_layout(state)

    page_header('🗺', 'Exploration Area', 'exploration_area')

    # Page-local state
    page: dict = {
        'selected_species': None,
    }

    # -----------------------------------------------------------------------
    # Right sidebar – species selection actions
    # -----------------------------------------------------------------------
    with ui.right_drawer(value=False).classes('bg-grey-1 p-4 w-64') as sidebar:
        ui.label('🎯 Selected Species').classes('text-h6 q-mb-sm')
        species_name_label = ui.label('–').classes('text-body1 font-bold q-mb-md')

        play_btn = ui.button(
            '▶ Play Species',
            on_click=lambda: _play_species(state, page),
        ).props('no-caps color=primary').classes('w-full q-mb-sm')

        xc_btn = ui.button('🔊 Xeno-Canto') \
            .props('no-caps flat').classes('w-full')
        xc_btn.disable()

    def _update_sidebar(scientific_name: Optional[str]) -> None:
        page['selected_species'] = scientific_name
        if scientific_name:
            species_name_label.set_text(scientific_name)
            xc_query = scientific_name.replace(' ', '+')
            xc_url = f'https://xeno-canto.org/explore?query={xc_query}&view=3'
            xc_btn.on('click', lambda: ui.navigate.to(xc_url, new_tab=True))
            xc_btn.enable()
            sidebar.show()
        else:
            species_name_label.set_text('–')
            xc_btn.disable()
            sidebar.hide()

    # -----------------------------------------------------------------------
    # Section 1: DB selection via DbFolderTree
    # -----------------------------------------------------------------------
    with section_card('📂', 'Database Selection', 'exploration_db_selection'):

        no_db_warning = ui.label(
            '⚠️ No database selected – please select a folder below.'
        ).classes('text-orange-7 q-mb-sm')

        # Track pending operations locally to avoid race with async temp_db_process
        _pending: dict = {'added': set(), 'removed': set()}

        def _on_db_selection_change(selected: Set[Path]) -> None:
            """Diff selected set vs. current source_dbs and send ADD/REMOVE ops."""
            current = _get_loaded_db_paths(state.active_db) | _pending['added']
            current -= _pending['removed']

            to_remove = current - selected
            to_add    = selected - current

            for p in to_remove:
                bundle.temp_db_queue.put({'op': 'remove', 'db_path': str(p / 'birdnet_analysis.db')})
                _pending['added'].discard(p)
                _pending['removed'].add(p)
            for p in to_add:
                bundle.temp_db_queue.put({'op': 'add', 'db_path': str(p / 'birdnet_analysis.db')})
                _pending['removed'].discard(p)
                _pending['added'].add(p)

            _refresh_no_db_warning(selected)

        def _refresh_no_db_warning(selected: Set[Path]) -> None:
            no_db_warning.set_visibility(len(selected) == 0)

        DbFolderTree(
            root_path=state.root_path,
            on_change=_on_db_selection_change,
            preselected=_get_loaded_db_paths(state.active_db),
        )
        # Set initial warning visibility based on current temp_db state
        _refresh_no_db_warning(_get_loaded_db_paths(state.active_db))

    # -----------------------------------------------------------------------
    # Section 2: Aggregate stats
    # -----------------------------------------------------------------------
    with section_card('📊', 'Database Overview', 'exploration_db_overview'):

        with ui.row().classes('gap-6 items-start flex-wrap') as stats_row:
            metric_dbs         = ui.column()
            metric_detections  = ui.column()
            metric_species     = ui.column()
            metric_files       = ui.column()
            metric_duration    = ui.column()

        no_stats_hint = ui.label(
            'No data yet – select databases above.'
        ).classes('text-caption text-grey-6')

        def _render_metric(container: ui.column, label: str, value: str) -> None:
            container.clear()
            with container:
                ui.label(label).classes('text-caption text-grey-6')
                ui.label(value).classes('text-body1 font-bold')

        @ui.refreshable
        def _render_stats() -> None:
            db = state.active_db
            if db is None:
                stats_row.set_visibility(False)
                no_stats_hint.set_visibility(True)
                return

            stats = _query_aggregate_stats(db)
            if stats['source_db_count'] == 0:
                stats_row.set_visibility(False)
                no_stats_hint.set_visibility(True)
                return

            stats_row.set_visibility(True)
            no_stats_hint.set_visibility(False)

            _render_metric(metric_dbs,        'Selected DBs',           str(stats['source_db_count']))
            _render_metric(metric_detections, 'Detections (total)',      str(stats['detection_count']))
            _render_metric(metric_species,    'Species',                 str(stats['species_count']))
            _render_metric(metric_files,      'Recording Files',         str(stats['file_count']))
            dur_h = stats['total_duration_s'] / 3600.0 if stats['total_duration_s'] else 0.0
            _render_metric(metric_duration,   'Total Duration',          f'{dur_h:.1f} h')

        _render_stats()

        # Timer refreshes stats while temp_db_process is running
        was_running = {'value': False}

        def _check_temp_db_running() -> None:
            tasks = bundle.shared_state.get('tasks', {})
            td = tasks.get('temp_db', {})
            running = td.get('running', False)
            if was_running['value'] and not running:
                # temp_db_process just finished an operation → refresh all
                _render_stats.refresh()
                _render_recording_files.refresh()
                _render_species_list.refresh()
            was_running['value'] = running

        ui.timer(2.0, _check_temp_db_running)

    # -----------------------------------------------------------------------
    # Section 3: Notes (disabled – single-DB management not yet implemented)
    # -----------------------------------------------------------------------
    with section_card('📝', 'Notes', 'exploration_notes'):
        ui.label(
            'Notes are temporarily disabled. '
            'Single-database management will be available in a future release.'
        ).classes('text-caption text-grey-6')
        ui.textarea(
            label='Database notes / comments',
            placeholder='Not available in aggregated view.',
        ).classes('w-full').props('rows=4 outlined readonly disable')

    # -----------------------------------------------------------------------
    # Section 4: Recording Files (all files across selected source DBs)
    # -----------------------------------------------------------------------
    with section_card('📁', 'Recording Files', 'exploration_recording_files'):
        _render_recording_files(state)

    # -----------------------------------------------------------------------
    # Section 5: Species List + Download
    # -----------------------------------------------------------------------
    with section_card('🦜', 'Species List', 'exploration_species_list'):
        _render_species_list(state, _update_sidebar)
        
# ===========================================================================
# Module-level helpers
# ===========================================================================

def _get_loaded_db_paths(temp_db: Optional[Path]) -> Set[Path]:
    """
    Return the set of source-DB *parent folders* currently in temp_db.

    source_dbs.db_path stores the absolute path to birdnet_analysis.db;
    we return the parent folder so it matches DbFolderTree.selected_folders.
    """
    if temp_db is None or not temp_db.exists():
        return set()
    try:
        conn = sqlite3.connect(temp_db)
        rows = conn.execute("SELECT db_path FROM source_dbs").fetchall()
        conn.close()
        return {Path(r[0]).parent for r in rows}
    except Exception as e:
        logger.error(f"_get_loaded_db_paths failed: {e}")
        return set()


def _query_aggregate_stats(temp_db: Path) -> dict:
    """
    Query aggregate statistics from temp_db.

    Returns dict with keys:
        source_db_count, detection_count, species_count,
        file_count, total_duration_s
    """
    defaults = {
        'source_db_count':  0,
        'detection_count':  0,
        'species_count':    0,
        'file_count':       0,
        'total_duration_s': 0.0,
    }
    try:
        conn = sqlite3.connect(temp_db)
        conn.row_factory = sqlite3.Row

        defaults['source_db_count'] = conn.execute(
            "SELECT COUNT(*) FROM source_dbs"
        ).fetchone()[0]

        defaults['detection_count'] = conn.execute(
            "SELECT COUNT(*) FROM detections"
        ).fetchone()[0]

        defaults['species_count'] = conn.execute(
            "SELECT COUNT(*) FROM species_list"
        ).fetchone()[0]

        defaults['file_count'] = conn.execute(
            "SELECT COUNT(*) FROM metadata"
        ).fetchone()[0]

        row = conn.execute(
            "SELECT SUM(duration_seconds) FROM metadata"
        ).fetchone()
        defaults['total_duration_s'] = row[0] or 0.0

        conn.close()
    except Exception as e:
        logger.error(f"_query_aggregate_stats failed: {e}")

    return defaults


@ui.refreshable
def _render_recording_files(state: AppState) -> None:
    """Render recording files table from temp_db metadata (all selected DBs)."""
    db = state.active_db
    if db is None or not db.exists():
        ui.label('No databases selected.').classes('text-caption text-grey-6')
        return

    try:
        conn = sqlite3.connect(db)
        conn.row_factory = sqlite3.Row
        rows_raw = conn.execute("""
            SELECT
                m.filename,
                m.timestamp_local,
                m.duration_seconds,
                m.temperature_c,
                m.battery_voltage,
                m.gps_lat,
                m.gps_lon,
                s.display_name
            FROM metadata m
            JOIN source_dbs s ON m.source_db_id = s.id
            ORDER BY m.timestamp_local ASC
        """).fetchall()
        conn.close()
    except Exception as e:
        logger.error(f"Recording files query failed: {e}")
        ui.label('Failed to load recording files.').classes('text-caption text-red-7')
        return

    if not rows_raw:
        ui.label('No recording files found.').classes('text-caption text-grey-6')
        return

    ui.label(f'Total files: {len(rows_raw)}').classes('text-caption q-mb-xs')

    rows = []
    for m in rows_raw:
        start = m['timestamp_local'] or '–'
        duration = m['duration_seconds'] or 0
        try:
            start_dt = datetime.fromisoformat(start)
            end_dt   = start_dt.replace(second=start_dt.second + int(duration))
            end_str  = end_dt.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            end_str = '–'
        gps = (
            f"{m['gps_lat']:.4f}, {m['gps_lon']:.4f}"
            if m['gps_lat'] and m['gps_lon'] else 'N/A'
        )
        rows.append({
            'DB':           m['display_name'],
            'Filename':     m['filename'],
            'Start':        start,
            'End':          end_str,
            'Duration (s)': f'{duration:.1f}',
            'Temp (°C)':    f"{m['temperature_c']:.1f}" if m['temperature_c'] else 'N/A',
            'Battery (V)':  f"{m['battery_voltage']:.2f}" if m['battery_voltage'] else 'N/A',
            'GPS':          gps,
        })

    ui.table(
        columns=[{'name': k, 'label': k, 'field': k} for k in rows[0]],
        rows=rows,
        pagination=20,
    ).classes('w-full').props('flat bordered dense')


@ui.refreshable
def _render_species_list(state: AppState, update_sidebar) -> None:
    """Render species AG Grid and download buttons from temp_db."""
    db = state.active_db
    if db is None or not db.exists():
        ui.label('No databases selected.').classes('text-caption text-grey-6')
        return

    if not species_list_exists(db):
        ui.label(
            '⚠️ Species list not yet available – waiting for database to load.'
        ).classes('text-caption text-orange-7')
        return

    labels = load_labels(state.bird_language_code)
    species_list = get_species_list_with_counts(db, labels=labels)

    if not species_list:
        ui.label('No species found.').classes('text-caption text-grey-6')
        return

    ui.label(f'Total species: {len(species_list)}').classes('text-caption q-mb-xs')

    min_score  = min(s['score'] for s in species_list)
    grid_rows: List[Dict] = [
        {
            'scientific_name': s['scientific_name'],
            'local_name':      s.get('local_name') or '',
            'detections':      format_detections_column(
                                   s['count_high'], s['count_low'],
                                   s['score'], min_score,
                               ),
            'score':           s['score'],
        }
        for s in species_list
    ]

    grid_cols = [
        {
            'headerName': 'Scientific Name', 'field': 'scientific_name',
            'width': 250, 'sortable': True, 'filter': True, 'resizable': True,
        },
        {
            'headerName': 'Local Name', 'field': 'local_name',
            'width': 250, 'sortable': True, 'filter': True, 'resizable': True,
        },
        {
            'headerName': 'Detections (≥70% / all) {score}', 'field': 'detections',
            'width': 260, 'sortable': False, 'filter': False, 'resizable': True,
        },
        {
            'headerName': 'Score', 'field': 'score',
            'hide': True, 'sort': 'desc',
        },
    ]

    grid = ui.aggrid({
        'columnDefs': grid_cols,
        'rowData':    grid_rows,
        'rowSelection': 'single',
        'defaultColDef': {'sortable': True, 'resizable': True},
    }).classes('w-full').style('height: 500px')

    async def _on_row_selected(e) -> None:
        rows_sel = await grid.get_selected_rows()
        if rows_sel:
            update_sidebar(rows_sel[0]['scientific_name'])
        else:
            update_sidebar(None)

    grid.on('rowSelected', _on_row_selected)

    ui.separator().classes('q-my-md')
    ui.label('📥 Download Species List').classes('text-subtitle1 q-mb-xs')

    download_scope = ui.radio(
        options={'all': 'All species', 'filtered': 'Only filtered species'},
        value='all',
    ).props('inline')

    with ui.row().classes('gap-3 q-mt-sm'):
        ui.button(
            '📥 Download CSV',
            on_click=lambda: asyncio.create_task(
                _download_csv(state, grid, download_scope.value, context.client)
            ),
        ).props('no-caps')
        ui.button(
            '📥 Download Excel',
            on_click=lambda: asyncio.create_task(
                _download_excel(state, grid, download_scope.value, context.client)
            ),
        ).props('no-caps')


def _play_species(state: AppState, page: dict) -> None:
    """Set audio player filter for selected species and navigate."""
    scientific = page.get('selected_species')
    if not scientific or state.active_db is None:
        return
    from ..db_queries import get_recording_date_range
    min_date, max_date = get_recording_date_range(state.active_db)
    state.ap_filter_species    = scientific
    state.ap_filter_date_from  = min_date.date() if min_date else None
    state.ap_filter_date_to    = max_date.date() if max_date else None
    state.ap_filter_use_time   = False
    state.ap_filter_confidence = None
    state.ap_filter_limit      = 25
    state.ap_filter_offset     = 0
    state.ap_filter_sort       = 'confidence'
    state.ap_filters_applied   = True
    ui.navigate.to('/audio-player')


async def _get_all_rows(state: AppState) -> List[Dict]:
    """Fetch all species rows for CSV/Excel export."""
    labels       = load_labels(state.bird_language_code)
    species_list = get_species_list_with_counts(state.active_db, labels=labels)
    if not species_list:
        return []
    min_score = min(s['score'] for s in species_list)
    return [
        {
            'Scientific Name': s['scientific_name'],
            'Local Name':      s.get('local_name') or '',
            'Detections':      format_detections_column(
                                   s['count_high'], s['count_low'],
                                   s['score'], min_score,
                               ),
        }
        for s in species_list
    ]


async def _download_csv(
    state: AppState, grid: ui.aggrid, scope: str, client
) -> None:
    with client:
        try:
            if scope == 'all':
                rows       = await _get_all_rows(state)
                scope_text = 'all'
            else:
                raw  = await grid.get_client_data(method='filtered_sorted', timeout=6.0)
                rows = [
                    {
                        'Scientific Name': r.get('scientific_name', ''),
                        'Local Name':      r.get('local_name', ''),
                        'Detections':      r.get('detections', ''),
                    }
                    for r in (raw or [])
                ]
                scope_text = 'filtered'

            buf = StringIO()
            buf.write(f'# Selections: {scope_text}\n')
            buf.write(f'# temp_db: {state.active_db}\n')
            buf.write('#\n')
            pd.DataFrame(rows).to_csv(buf, index=False)
            ui.download(buf.getvalue().encode('utf-8'), 'species_export.csv')

        except Exception as e:
            logger.error(f'CSV download failed: {e}')
            ui.notify(f'CSV export failed: {e}', type='negative')


async def _download_excel(
    state: AppState, grid: ui.aggrid, scope: str, client
) -> None:
    with client:
        try:
            from openpyxl.styles import Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            if scope == 'all':
                rows       = await _get_all_rows(state)
                scope_text = 'all'
            else:
                raw  = await grid.get_client_data(method='filtered_sorted', timeout=6.0)
                rows = [
                    {
                        'Scientific Name': r.get('scientific_name', ''),
                        'Local Name':      r.get('local_name', ''),
                        'Detections':      r.get('detections', ''),
                    }
                    for r in (raw or [])
                ]
                scope_text = 'filtered'

            df  = pd.DataFrame(rows)
            buf = BytesIO()
            thin = Border(bottom=Side(style='thin'))

            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                info_lines = [
                    f'Selections: {scope_text}',
                    f'temp_db: {state.active_db}',
                ]
                pd.DataFrame({'Info': info_lines}).to_excel(
                    writer, sheet_name='Info', index=False, header=False
                )
                ws_info = writer.sheets['Info']
                ws_info.column_dimensions['A'].width = 80
                for row in ws_info.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                df.to_excel(writer, sheet_name='Species Table', index=False)
                ws = writer.sheets['Species Table']
                for i, w in enumerate([22, 25, 20, 28], start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = thin
                ws.oddFooter.right.text = 'Page &P'

            ui.download(buf.getvalue(), 'species_export.xlsx')

        except Exception as e:
            logger.error(f'Excel download failed: {e}')
            ui.notify(f'Excel export failed: {e}', type='negative')