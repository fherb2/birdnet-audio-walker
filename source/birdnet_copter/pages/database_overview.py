"""
Database Overview page.
Covers: page route, layout, folder tree, DB info, actualize button, notes field,
        recording files table, species AG Grid, downloads, right sidebar.
"""

import asyncio
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from loguru import logger
from nicegui import ui, app as nicegui_app, context

from ..app_state import AppState
from ..pages.layout import create_layout
from ..gui_elements.folder_tree import FolderTree
from shared.db_queries import (
    get_analysis_config,
    set_analysis_config,
    get_species_count,
    get_species_list_with_counts,
    format_detections_column,
    create_species_list_table,
    get_all_metadata,
    get_recording_date_range,
)
from ..bird_language import load_labels


# ---------------------------------------------------------------------------
# Page registration
# ---------------------------------------------------------------------------

def _get_state() -> AppState:
    return nicegui_app.state.app_state  # type: ignore[attr-defined]


@ui.page('/')
async def database_overview() -> None:
    state = _get_state()
    create_layout(state)

    # -----------------------------------------------------------------------
    # Page-local reactive variables
    # -----------------------------------------------------------------------
    page: Dict = {
        'selected_folder': state.active_db.parent if state.active_db else None,
        'actualize_running': False,
        'selected_species': None,
    }

    # -----------------------------------------------------------------------
    # Right sidebar – species selection actions
    # -----------------------------------------------------------------------
    with ui.right_drawer(value=False).classes('bg-grey-1 p-4 w-64') as sidebar:
        ui.label('🎯 Selected Species').classes('text-h6 q-mb-sm')
        species_name_label = ui.label('–').classes('text-body1 font-bold q-mb-md')

        play_btn = ui.button(
            '▶ Play Species',
            on_click=lambda: _play_species(state, page),
        ).props('no-caps color=primary').classes('w-full q-mb-sm')

        xc_btn = ui.button('🔊 Xeno-Canto') \
            .props('no-caps flat').classes('w-full')
        xc_btn.disable()

    def _update_sidebar(scientific_name: Optional[str]) -> None:
        page['selected_species'] = scientific_name
        if scientific_name:
            species_name_label.set_text(scientific_name)
            xc_query = scientific_name.replace(' ', '+')
            xc_url = f'https://xeno-canto.org/explore?query={xc_query}&view=3'
            xc_btn.on('click', lambda: ui.navigate.to(xc_url, new_tab=True))
            xc_btn.enable()
            sidebar.show()
        else:
            species_name_label.set_text('–')
            xc_btn.disable()
            sidebar.hide()

    # -----------------------------------------------------------------------
    # Section: Folder tree
    # -----------------------------------------------------------------------
    ui.label('📂 Database Selection').classes('text-h6 q-mt-md q-mb-xs')

    def on_folder_selected(folder: Path) -> None:
        page['selected_folder'] = folder
        db_path = folder / 'birdnet_analysis.db'
        if db_path.exists():
            if db_path != state.active_db:
                state.reset_filter_state()
                state.active_db = db_path
                logger.info(f"Switched to database: {db_path}")
            _refresh_db_info()
            _render_recording_files.refresh()
            _render_species_list.refresh()
            _update_sidebar(None)
        else:
            state.active_db = None
            _refresh_db_info()
            _render_recording_files.refresh()
            _render_species_list.refresh()
            _update_sidebar(None)

    def on_root_changed(new_root: Path) -> None:
        state.root_path = new_root
        logger.info(f"Tree root changed to: {new_root}")

    FolderTree(
        root_path=state.root_path,
        on_select=on_folder_selected,
        on_root_change=on_root_changed,
    )

    ui.separator().classes('q-my-md')

    # -----------------------------------------------------------------------
    # Section: Database info
    # -----------------------------------------------------------------------
    ui.label('📊 Database Information').classes('text-h6 q-mb-xs')

    with ui.row().classes('gap-4 items-start') as db_info_row:
        metric_conf    = ui.column()
        metric_created = ui.column()
        metric_species = ui.column()

    no_db_hint = ui.label('No database selected.').classes('text-caption text-grey-6')

    def _render_metric(container: ui.column, label: str, value: str) -> None:
        container.clear()
        with container:
            ui.label(label).classes('text-caption text-grey-6')
            ui.label(value).classes('text-body1 font-bold')

    def _refresh_db_info() -> None:
        db = state.active_db
        if db is None or not db.exists():
            db_info_row.set_visibility(False)
            no_db_hint.set_visibility(True)
            actualize_row.set_visibility(False)
            notes_section.set_visibility(False)
            return

        db_info_row.set_visibility(True)
        no_db_hint.set_visibility(False)
        actualize_row.set_visibility(True)
        notes_section.set_visibility(True)

        conf     = get_analysis_config(db, 'confidence_threshold')
        conf_str = f"{float(conf)*100:.0f}%" if conf else '–'
        created  = get_analysis_config(db, 'created_at') or '–'
        count    = get_species_count(db)
        count_str = str(count) if count > 0 else 'not available'

        _render_metric(metric_conf,    'Confidence Threshold', conf_str)
        _render_metric(metric_created, 'Created',              created)
        _render_metric(metric_species, 'Species',              count_str)

        comment = get_analysis_config(db, 'user_comment') or ''
        notes_area.set_value(comment)

    ui.separator().classes('q-my-md')

    # -----------------------------------------------------------------------
    # Section: Actualize species list
    # -----------------------------------------------------------------------
    with ui.row().classes('items-center gap-3') as actualize_row:
        act_btn = ui.button(
            '🔄 Actualize Species List',
            on_click=lambda: asyncio.create_task(_run_actualize()),
        ).props('no-caps')
        if state.read_only:
            act_btn.disable()
        act_spinner = ui.spinner(size='sm')
        act_spinner.set_visibility(False)
        act_status = ui.label('').classes('text-caption')

    async def _run_actualize() -> None:
        if page['actualize_running'] or state.active_db is None:
            return
        page['actualize_running'] = True
        act_btn.disable()
        act_spinner.set_visibility(True)
        act_status.set_text('Updating…')
        try:
            loop = asyncio.get_event_loop()
            success = await loop.run_in_executor(
                None, create_species_list_table, state.active_db
            )
            if success:
                count = get_species_count(state.active_db)
                act_status.set_text(f'✅ Done – {count} species')
                _refresh_db_info()
            else:
                act_status.set_text('❌ Failed')
        except Exception as e:
            logger.error(f"Actualize failed: {e}")
            act_status.set_text(f'❌ Error: {e}')
        finally:
            page['actualize_running'] = False
            act_btn.enable()
            act_spinner.set_visibility(False)

    ui.separator().classes('q-my-md')

    # -----------------------------------------------------------------------
    # Section: Notes
    # -----------------------------------------------------------------------
    with ui.column().classes('w-full gap-2') as notes_section:
        ui.label('📝 Notes').classes('text-h6')
        notes_area = ui.textarea(
            label='Database notes / comments',
            placeholder='Add notes about this recording session…',
        ).classes('w-full').props('rows=6 outlined')
        if state.read_only:
            notes_area.props('readonly')

        save_btn = ui.button(
            '💾 Save Notes',
            on_click=lambda: _save_notes(),
        ).props('no-caps')
        if state.read_only:
            save_btn.disable()
        save_status = ui.label('').classes('text-caption')

    def _save_notes() -> None:
        if state.active_db is None:
            return
        ok = set_analysis_config(state.active_db, 'user_comment', notes_area.value)
        save_status.set_text('✅ Saved' if ok else '❌ Failed to save')

    # Initial render
    _refresh_db_info()
    if state.active_db is None or not state.active_db.exists():
        actualize_row.set_visibility(False)
        notes_section.set_visibility(False)
    else:
        comment = get_analysis_config(state.active_db, 'user_comment') or ''
        notes_area.set_value(comment)

    ui.separator().classes('q-my-md')

    # -----------------------------------------------------------------------
    # Section: Recording files table
    # -----------------------------------------------------------------------
    @ui.refreshable
    def _render_recording_files():
        ui.label('📁 Recording Files').classes('text-h6 q-mb-xs')

        if state.active_db is None or not state.active_db.exists():
            ui.label('No database selected.').classes('text-caption text-grey-6')
            return

        metadata_list = get_all_metadata(state.active_db)
        if not metadata_list:
            ui.label('No recording files found.').classes('text-caption text-grey-6')
            return

        ui.label(f'Total files: {len(metadata_list)}').classes('text-caption')
        rows = []
        for m in metadata_list:
            start = m['timestamp_local'] or '–'
            duration = m['duration_seconds'] or 0
            try:
                start_dt = datetime.fromisoformat(start)
                end_dt = start_dt.replace(
                    second=start_dt.second + int(duration)
                )
                end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                end_str = '–'
            gps = (
                f"{m['gps_lat']:.4f}, {m['gps_lon']:.4f}"
                if m.get('gps_lat') and m.get('gps_lon')
                else 'N/A'
            )
            rows.append({
                'Filename':     m['filename'],
                'Start':        start,
                'End':          end_str,
                'Duration (s)': f"{duration:.1f}",
                'Temp (°C)':    f"{m['temperature_c']:.1f}" if m.get('temperature_c') else 'N/A',
                'Battery (V)':  f"{m['battery_voltage']:.2f}" if m.get('battery_voltage') else 'N/A',
                'GPS':          gps,
            })
        ui.table(
            columns=[{'name': k, 'label': k, 'field': k} for k in rows[0]],
            rows=rows,
            pagination=10,
        ).classes('w-full').props('flat bordered dense')

    # -----------------------------------------------------------------------
    # Section: Species List
    # -----------------------------------------------------------------------
    @ui.refreshable
    def _render_species_list():
        ui.label('🦜 Species List').classes('text-h6 q-mb-xs')

        if state.active_db is None or not state.active_db.exists():
            ui.label('No database selected.').classes('text-caption text-grey-6')
            return

        labels = load_labels(state.bird_language_code)
        species_list = get_species_list_with_counts(state.active_db, labels=labels)
        if not species_list:
            ui.label(
                '⚠️ Species list not available. '
                'Click "Actualize Species List" above.'
            ).classes('text-caption text-orange-7')
            return

        ui.label(f'Total species: {len(species_list)}').classes('text-caption')

        min_score = min(s['score'] for s in species_list)
        grid_rows: List[Dict] = []
        for s in species_list:
            grid_rows.append({
                'scientific_name': s['scientific_name'],
                'local_name':      s.get('local_name') or '',
                'detections':      format_detections_column(
                                       s['count_high'], s['count_low'],
                                       s['score'], min_score,
                                   ),
                'score':           s['score'],
            })

        grid_cols = [
            {
                'headerName': 'Scientific Name', 'field': 'scientific_name',
                'width': 250, 'sortable': True, 'filter': True, 'resizable': True,
            },
            {
                'headerName': 'Local Name', 'field': 'local_name',
                'width': 250, 'sortable': True, 'filter': True, 'resizable': True,
            },
            {
                'headerName': 'Detections (≥70% / all) {score}', 'field': 'detections',
                'width': 260, 'sortable': False, 'filter': False, 'resizable': True,
            },
            {
                'headerName': 'Score', 'field': 'score',
                'hide': True, 'sort': 'desc',
            },
        ]
        
        grid = ui.aggrid({
            'columnDefs': grid_cols,
            'rowData': grid_rows,
            'rowSelection': 'single',
            'defaultColDef': {'sortable': True, 'resizable': True},
        }).classes('w-full').style('height: 500px')

        async def on_row_selected(e) -> None:
            rows_sel = await grid.get_selected_rows()
            if rows_sel:
                _update_sidebar(rows_sel[0]['scientific_name'])
            else:
                _update_sidebar(None)

        grid.on('rowSelected', on_row_selected)

        ui.separator().classes('q-my-md')

        ui.label('📥 Download Species List').classes('text-h6 q-mb-xs')

        download_scope = ui.radio(
            options={'all': 'All species', 'filtered': 'Only filtered species'},
            value='all',
        ).props('inline')

        with ui.row().classes('gap-3 q-mt-sm'):
            ui.button(
                '📥 Download CSV',
                on_click=lambda: _download_csv(state, grid, download_scope.value, context.client),
            ).props('no-caps')
            ui.button(
                '📥 Download Excel',
                on_click=lambda: _download_excel(state, grid, download_scope.value, context.client),
            ).props('no-caps')

    ui.separator().classes('q-my-md')
    _render_recording_files()
    ui.separator().classes('q-my-md')
    _render_species_list()


# ===========================================================================
# Module-level helpers (outside page handler)
# ===========================================================================

def _play_species(state: AppState, page: dict) -> None:
    """Set audio player filter for selected species and navigate."""
    scientific = page.get('selected_species')
    if not scientific or state.active_db is None:
        return
    min_date, max_date = get_recording_date_range(state.active_db)
    state.ap_filter_species    = scientific
    state.ap_filter_date_from  = min_date.date() if min_date else None
    state.ap_filter_date_to    = max_date.date() if max_date else None
    state.ap_filter_use_time   = False
    state.ap_filter_confidence = None
    state.ap_filter_limit      = 25
    state.ap_filter_offset     = 0
    state.ap_filter_sort       = 'confidence'
    state.ap_filters_applied   = True
    ui.navigate.to('/audio-player')


def _build_meta_header(state: AppState) -> dict:
    db = state.active_db
    min_d, max_d = get_recording_date_range(db) if db else (None, None)
    return {
        'db_path':   str(db) if db else '–',
        'date_from': min_d.strftime('%Y-%m-%d') if min_d else '–',
        'date_to':   max_d.strftime('%Y-%m-%d') if max_d else '–',
        'notes':     get_analysis_config(db, 'user_comment') or '' if db else '',
    }


async def _get_all_rows(state: AppState) -> List[Dict]:
    labels = load_labels(state.bird_language_code)
    species_list = get_species_list_with_counts(state.active_db, labels=labels)
    if not species_list:
        return []
    min_score = min(s['score'] for s in species_list)
    return [
        {
            'Scientific Name': s['scientific_name'],
            'Local Name':      s.get('local_name') or '',
            'Detections':      format_detections_column(
                                   s['count_high'], s['count_low'],
                                   s['score'], min_score,
                               ),
        }
        for s in species_list
    ]


async def _download_csv(state: AppState, grid: ui.aggrid, scope: str, client) -> None:
    with client:
        try:
            meta = _build_meta_header(state)
            if scope == 'all':
                rows = await _get_all_rows(state)
                scope_text = 'all'
            else:
                raw = await grid.get_client_data(method='filtered_sorted', timeout=6.0)
                rows = [
                    {
                        'Scientific Name': r.get('scientific_name', ''),
                        'Local Name':      r.get('local_name', ''),
                        'Detections':      r.get('detections', ''),
                    }
                    for r in (raw or [])
                ]
                scope_text = 'filtered'

            buf = StringIO()
            buf.write(f"# File: {meta['db_path']}\n")
            buf.write(f"# Selections: {scope_text}\n")
            buf.write(f"# From: {meta['date_from']}  To: {meta['date_to']}\n")
            buf.write("#\n")
            if meta['notes']:
                for line in meta['notes'].splitlines():
                    buf.write(f"# {line}\n")
                buf.write("#\n")
            pd.DataFrame(rows).to_csv(buf, index=False)
            
            db_stem = state.active_db.parent.name if state.active_db else 'export'
            ui.download(buf.getvalue().encode('utf-8'), f'species_{db_stem}.csv')

        except Exception as e:
            logger.error(f"CSV download failed: {e}")
            ui.notify(f'CSV export failed: {e}', type='negative')


async def _download_excel(state: AppState, grid: ui.aggrid, scope: str, client) -> None:
    with client:
        try:
            from openpyxl.styles import Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            meta = _build_meta_header(state)
            if scope == 'all':
                rows = await _get_all_rows(state)
                scope_text = 'all'
            else:
                raw = await grid.get_client_data(method='filtered_sorted', timeout=6.0)
                rows = [
                    {
                        'Scientific Name': r.get('scientific_name', ''),
                        'Local Name':      r.get('local_name', ''),
                        'Detections':      r.get('detections', ''),
                    }
                    for r in (raw or [])
                ]
                scope_text = 'filtered'

            df = pd.DataFrame(rows)
            buf = BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                info_lines = [
                    f"File: {meta['db_path']}",
                    f"Selections: {scope_text}",
                    f"From: {meta['date_from']}  To: {meta['date_to']}",
                    '', 'Notes:',
                ] + meta['notes'].splitlines()
                pd.DataFrame({'Info': info_lines}).to_excel(
                    writer, sheet_name='Database', index=False, header=False
                )
                ws_info = writer.sheets['Database']
                ws_info.column_dimensions['A'].width = 80
                thin = Border(bottom=Side(style='thin'))
                for row in ws_info.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True)

                df.to_excel(writer, sheet_name='Species Table', index=False)
                ws = writer.sheets['Species Table']
                for i, w in enumerate([22, 25, 20, 28], start=1):
                    ws.column_dimensions[get_column_letter(i)].width = w
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = thin
                ws.oddHeader.left.text = str(meta['db_path'])
                ws.oddFooter.right.text = 'Page &P'

            db_stem = state.active_db.parent.name if state.active_db else 'export'
            ui.download(buf.getvalue(), f'species_{db_stem}.xlsx')

        except Exception as e:
            logger.error(f"Excel download failed: {e}")
            ui.notify(f'Excel export failed: {e}', type='negative')